"""
MARKET SPACE - SAAS FINANCIAL MODEL GENERATOR
==================================================

This script generates a comprehensive 5-year financial model for Market Space
suitable for investor presentations.

SETUP INSTRUCTIONS:
-------------------
1. Install dependencies:
   pip install openpyxl pandas numpy

2. (Optional) Create a business-model-canvas.md file with your business details

3. Run the script:
   python financial-model.py

4. Output: saas_financial_model.xlsx

CUSTOMIZATION:
--------------
- Edit the ASSUMPTIONS dictionary below to adjust your business parameters
- If you have a business-model-canvas.md file, the script will extract values from it
- All assumptions appear in a dedicated "Assumptions" sheet in the Excel file
- Use color coding: BLUE = inputs, BLACK = calculations, GREEN = revenue, RED = costs

KEY FORMULAS:
-------------
- MRR = Sum of all monthly subscription revenue
- ARR = MRR × 12
- CAC = Total Sales & Marketing Spend / New Customers Acquired
- LTV = (ARPU × Gross Margin %) / Monthly Churn Rate
- LTV:CAC Ratio = LTV / CAC (Target: 3:1 or higher)
- Magic Number = (Net New ARR / Sales & Marketing Spend) (Target: 0.75+)
- Rule of 40 = Revenue Growth % + EBITDA Margin % (Target: 40%+)
- NRR = (Starting MRR + Expansion - Churn) / Starting MRR

AUTHOR: Claude Code
VERSION: 1.0
DATE: 2025
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import os
from datetime import datetime
from decimal import Decimal
import json

# =============================================================================
# BUSINESS MODEL CANVAS PARSER
# =============================================================================

def parse_business_model_canvas(filepath='business-model-canvas.md'):
    """
    Parse business model canvas markdown file and extract financial assumptions
    """
    assumptions = {}

    if not os.path.exists(filepath):
        print(f"[!] Business Model Canvas file not found at {filepath}")
        print("[*] Using default assumptions for Market Space marketplace model\n")
        return None

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()

        # Extract customer segments for pricing tiers
        # Extract revenue streams
        # Extract cost structure
        # etc.

        print(f"[OK] Loaded business context from {filepath}\n")
        return assumptions
    except Exception as e:
        print(f"[!] Error reading canvas file: {e}")
        return None

# =============================================================================
# FINANCIAL ASSUMPTIONS
# =============================================================================

# These can be overridden by business-model-canvas.md or manually edited
ASSUMPTIONS = {
    # BUSINESS INFORMATION
    'company_name': 'Market Space',
    'model_date': datetime.now().strftime('%Y-%m-%d'),
    'currency': 'USD',

    # PRICING TIERS (Marketplace/SaaS Hybrid Model)
    'pricing_tiers': {
        'host_commission': {
            'name': 'Host Transaction Fee',
            'price_monthly': 0,  # Commission-based
            'commission_rate': 0.10,  # 10% of transaction
            'description': 'Commission on successful bookings'
        },
        'renter_fee': {
            'name': 'Renter Service Fee',
            'price_monthly': 0,
            'commission_rate': 0.05,  # 5% of transaction
            'description': 'Service fee on bookings'
        },
        'premium_listing': {
            'name': 'Premium Listing Subscription',
            'price_monthly': 29.99,
            'commission_rate': 0,
            'description': 'Enhanced visibility for hosts'
        },
        'pro_host': {
            'name': 'Pro Host Plan',
            'price_monthly': 99.99,
            'commission_rate': 0.08,  # Reduced commission
            'description': 'Advanced tools + lower commission'
        }
    },

    # CUSTOMER ACQUISITION
    'starting_customers': {
        'month_1': {
            'hosts_free': 10,
            'hosts_premium': 2,
            'hosts_pro': 0,
            'renters': 15
        }
    },

    # GROWTH ASSUMPTIONS
    'monthly_growth_rate_y1': 0.15,  # 15% MoM growth Year 1
    'monthly_growth_rate_y2': 0.10,  # 10% MoM growth Year 2
    'monthly_growth_rate_y3': 0.07,  # 7% MoM growth Year 3
    'monthly_growth_rate_y4': 0.05,  # 5% MoM growth Year 4-5

    # CONVERSION RATES
    'free_to_premium_rate': 0.08,  # 8% of free hosts convert monthly
    'premium_to_pro_rate': 0.05,  # 5% of premium convert to pro
    'renter_activation_rate': 0.60,  # 60% of renters make a booking

    # TRANSACTION ASSUMPTIONS
    'avg_transaction_value': 2500,  # Average booking value
    'transactions_per_renter_per_month': 0.25,  # 1 booking per 4 months

    # CHURN RATES (Monthly)
    'churn_rate_free': 0.10,  # 10% monthly churn
    'churn_rate_premium': 0.05,  # 5% monthly churn
    'churn_rate_pro': 0.03,  # 3% monthly churn
    'churn_rate_renter': 0.08,

    # EXPANSION REVENUE
    'upsell_rate': 0.03,  # 3% upgrade per month
    'cross_sell_rate': 0.02,  # 2% buy additional services
    'avg_expansion_value': 20,  # $20 average expansion

    # COST OF REVENUE (% of revenue)
    'hosting_infrastructure': 0.08,  # AWS, servers
    'payment_processing': 0.029,  # Stripe fees
    'customer_support': 0.12,  # Support team
    'total_cogs_percent': 0.23,  # Target 77% gross margin

    # SALES & MARKETING
    'cac_by_channel': {
        'paid_search': 150,
        'content_marketing': 75,
        'referral': 25,
        'partnerships': 100,
        'direct_sales': 300
    },
    'channel_mix': {
        'paid_search': 0.30,
        'content_marketing': 0.25,
        'referral': 0.20,
        'partnerships': 0.15,
        'direct_sales': 0.10
    },
    'marketing_spend_y1_monthly': 15000,
    'marketing_growth_rate': 1.20,  # 20% increase per year

    # RESEARCH & DEVELOPMENT
    'rd_percent_of_revenue': 0.25,  # 25% in early stage
    'rd_min_monthly': 20000,  # Minimum R&D spend

    # GENERAL & ADMINISTRATIVE
    'ga_percent_of_revenue': 0.15,
    'ga_min_monthly': 10000,

    # HEADCOUNT PLANNING
    'initial_headcount': {
        'engineering': 3,
        'sales': 2,
        'marketing': 1,
        'operations': 1,
        'executive': 2
    },
    'avg_salary': {
        'engineering': 120000,
        'sales': 80000,
        'marketing': 75000,
        'operations': 65000,
        'executive': 150000
    },
    'benefits_burden': 0.25,  # 25% of salary
    'new_hires_schedule': {
        'Q1_Y1': 0, 'Q2_Y1': 1, 'Q3_Y1': 2, 'Q4_Y1': 2,
        'Y2': 8, 'Y3': 12, 'Y4': 8, 'Y5': 5
    },

    # CASH & FINANCING
    'starting_cash': 500000,
    'seed_round': {'timing': 'M6', 'amount': 1000000, 'dilution': 0.20},
    'series_a': {'timing': 'M18', 'amount': 5000000, 'dilution': 0.25},

    # BENCHMARKS & TARGETS
    'target_ltv_cac_ratio': 3.0,
    'target_gross_margin': 0.75,
    'target_magic_number': 0.75,
    'target_rule_of_40': 40,
    'target_payback_months': 12
}

# =============================================================================
# EXCEL STYLING
# =============================================================================

class ExcelStyles:
    """Consistent styling for professional Excel output"""

    # Fonts
    HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    SUBHEADER_FONT = Font(name='Calibri', size=12, bold=True)
    TITLE_FONT = Font(name='Calibri', size=18, bold=True, color='1F4E78')
    INPUT_FONT = Font(name='Calibri', size=10, color='0000FF')

    # Fills
    HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    INPUT_FILL = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    REVENUE_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    COST_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    METRIC_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    # Alignment
    CENTER = Alignment(horizontal='center', vertical='center')
    RIGHT = Alignment(horizontal='right', vertical='center')
    LEFT = Alignment(horizontal='left', vertical='center')

    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# =============================================================================
# FINANCIAL CALCULATIONS ENGINE
# =============================================================================

class SaaSFinancialModel:
    def __init__(self, assumptions):
        self.assumptions = assumptions
        self.periods = 60  # 5 years monthly
        self.results = {}

    def calculate_revenue_model(self):
        """Calculate detailed revenue projections"""
        revenue = {
            'months': list(range(1, self.periods + 1)),
            'host_commission': [],
            'renter_fees': [],
            'premium_subscriptions': [],
            'pro_subscriptions': [],
            'expansion_revenue': [],
            'total_mrr': [],
            'arr': [],
            'customers_by_tier': {
                'hosts_free': [],
                'hosts_premium': [],
                'hosts_pro': [],
                'renters_active': []
            }
        }

        # Initialize starting customers
        hosts_free = self.assumptions['starting_customers']['month_1']['hosts_free']
        hosts_premium = self.assumptions['starting_customers']['month_1']['hosts_premium']
        hosts_pro = self.assumptions['starting_customers']['month_1']['hosts_pro']
        renters = self.assumptions['starting_customers']['month_1']['renters']

        for month in range(self.periods):
            # Determine growth rate based on year
            year = month // 12
            if year == 0:
                growth_rate = self.assumptions['monthly_growth_rate_y1']
            elif year == 1:
                growth_rate = self.assumptions['monthly_growth_rate_y2']
            elif year == 2:
                growth_rate = self.assumptions['monthly_growth_rate_y3']
            else:
                growth_rate = self.assumptions['monthly_growth_rate_y4']

            # Apply growth
            hosts_free *= (1 + growth_rate)
            renters *= (1 + growth_rate)

            # Apply conversions
            conversions_to_premium = hosts_free * self.assumptions['free_to_premium_rate']
            hosts_free -= conversions_to_premium
            hosts_premium += conversions_to_premium

            conversions_to_pro = hosts_premium * self.assumptions['premium_to_pro_rate']
            hosts_premium -= conversions_to_pro
            hosts_pro += conversions_to_pro

            # Apply churn
            hosts_free *= (1 - self.assumptions['churn_rate_free'])
            hosts_premium *= (1 - self.assumptions['churn_rate_premium'])
            hosts_pro *= (1 - self.assumptions['churn_rate_pro'])
            renters *= (1 - self.assumptions['churn_rate_renter'])

            # Calculate transactions
            active_renters = renters * self.assumptions['renter_activation_rate']
            monthly_transactions = active_renters * self.assumptions['transactions_per_renter_per_month']
            transaction_volume = monthly_transactions * self.assumptions['avg_transaction_value']

            # Calculate revenue components
            host_commission = transaction_volume * self.assumptions['pricing_tiers']['host_commission']['commission_rate']
            renter_fee = transaction_volume * self.assumptions['pricing_tiers']['renter_fee']['commission_rate']
            premium_sub = hosts_premium * self.assumptions['pricing_tiers']['premium_listing']['price_monthly']
            pro_sub = hosts_pro * self.assumptions['pricing_tiers']['pro_host']['price_monthly']

            # Expansion revenue (upsells, cross-sells)
            total_paying = hosts_premium + hosts_pro
            expansion = total_paying * self.assumptions['upsell_rate'] * self.assumptions['avg_expansion_value']

            # Store results
            revenue['host_commission'].append(host_commission)
            revenue['renter_fees'].append(renter_fee)
            revenue['premium_subscriptions'].append(premium_sub)
            revenue['pro_subscriptions'].append(pro_sub)
            revenue['expansion_revenue'].append(expansion)

            month_mrr = host_commission + renter_fee + premium_sub + pro_sub + expansion
            revenue['total_mrr'].append(month_mrr)
            revenue['arr'].append(month_mrr * 12)

            revenue['customers_by_tier']['hosts_free'].append(int(hosts_free))
            revenue['customers_by_tier']['hosts_premium'].append(int(hosts_premium))
            revenue['customers_by_tier']['hosts_pro'].append(int(hosts_pro))
            revenue['customers_by_tier']['renters_active'].append(int(active_renters))

        self.results['revenue'] = revenue
        return revenue

    def calculate_expenses(self):
        """Calculate detailed expense projections"""
        revenue = self.results['revenue']

        expenses = {
            'months': list(range(1, self.periods + 1)),
            'cogs': [],
            'sales_marketing': [],
            'research_development': [],
            'general_admin': [],
            'total_opex': [],
            'headcount': [],
            'new_customers': []
        }

        headcount = sum(self.assumptions['initial_headcount'].values())

        for month in range(self.periods):
            mrr = revenue['total_mrr'][month]

            # COGS (% of revenue)
            cogs = mrr * self.assumptions['total_cogs_percent']

            # Sales & Marketing (based on new customer acquisition)
            if month == 0:
                new_customers = sum(self.assumptions['starting_customers']['month_1'].values())
            else:
                prev_customers = sum([
                    revenue['customers_by_tier']['hosts_premium'][month-1],
                    revenue['customers_by_tier']['hosts_pro'][month-1],
                    revenue['customers_by_tier']['renters_active'][month-1]
                ])
                curr_customers = sum([
                    revenue['customers_by_tier']['hosts_premium'][month],
                    revenue['customers_by_tier']['hosts_pro'][month],
                    revenue['customers_by_tier']['renters_active'][month]
                ])
                new_customers = max(0, curr_customers - prev_customers)

            # Weighted average CAC based on channel mix
            blended_cac = sum([
                self.assumptions['cac_by_channel'][channel] * weight
                for channel, weight in self.assumptions['channel_mix'].items()
            ])

            year = month // 12
            marketing_budget = self.assumptions['marketing_spend_y1_monthly'] * (
                self.assumptions['marketing_growth_rate'] ** year
            )
            customer_acq_cost = new_customers * blended_cac
            sales_marketing = max(marketing_budget, customer_acq_cost)

            # R&D
            rd_calculated = mrr * self.assumptions['rd_percent_of_revenue']
            rd = max(rd_calculated, self.assumptions['rd_min_monthly'])

            # G&A
            ga_calculated = mrr * self.assumptions['ga_percent_of_revenue']
            ga = max(ga_calculated, self.assumptions['ga_min_monthly'])

            # Add headcount expansion
            quarter = month // 3
            year = month // 12
            if month % 3 == 0:  # First month of quarter
                if year == 0:
                    quarter_key = f'Q{(quarter % 4) + 1}_Y1'
                    if quarter_key in self.assumptions['new_hires_schedule']:
                        headcount += self.assumptions['new_hires_schedule'][quarter_key]
                elif year in [1, 2, 3, 4]:
                    year_key = f'Y{year + 1}'
                    if year_key in self.assumptions['new_hires_schedule'] and month % 12 == 0:
                        headcount += self.assumptions['new_hires_schedule'][year_key] / 4

            expenses['cogs'].append(cogs)
            expenses['sales_marketing'].append(sales_marketing)
            expenses['research_development'].append(rd)
            expenses['general_admin'].append(ga)
            expenses['total_opex'].append(cogs + sales_marketing + rd + ga)
            expenses['headcount'].append(int(headcount))
            expenses['new_customers'].append(int(new_customers))

        self.results['expenses'] = expenses
        return expenses

    def calculate_cash_flow(self):
        """Calculate cash flow and funding events"""
        revenue = self.results['revenue']
        expenses = self.results['expenses']

        cash_flow = {
            'months': list(range(1, self.periods + 1)),
            'revenue': revenue['total_mrr'],
            'total_expenses': expenses['total_opex'],
            'net_income': [],
            'cash_balance': [],
            'burn_rate': [],
            'runway_months': [],
            'funding_events': []
        }

        cash_balance = self.assumptions['starting_cash']

        for month in range(self.periods):
            net_income = revenue['total_mrr'][month] - expenses['total_opex'][month]
            cash_balance += net_income

            # Check for funding events
            funding = 0
            if month == 5 and self.assumptions['seed_round']:  # Month 6
                funding = self.assumptions['seed_round']['amount']
                cash_balance += funding
            elif month == 17 and self.assumptions['series_a']:  # Month 18
                funding = self.assumptions['series_a']['amount']
                cash_balance += funding

            # Calculate burn rate (negative net income)
            burn = min(0, net_income)

            # Calculate runway
            if burn < 0 and cash_balance > 0:
                runway = cash_balance / abs(burn)
            else:
                runway = 999  # Infinite runway (profitable)

            cash_flow['net_income'].append(net_income)
            cash_flow['cash_balance'].append(cash_balance)
            cash_flow['burn_rate'].append(burn)
            cash_flow['runway_months'].append(runway)
            cash_flow['funding_events'].append(funding)

        self.results['cash_flow'] = cash_flow
        return cash_flow

    def calculate_metrics(self):
        """Calculate SaaS-specific KPIs"""
        revenue = self.results['revenue']
        expenses = self.results['expenses']
        cash_flow = self.results['cash_flow']

        metrics = {
            'months': list(range(1, self.periods + 1)),
            'arr': revenue['arr'],
            'mrr': revenue['total_mrr'],
            'customers_total': [],
            'arpu': [],
            'cac': [],
            'ltv': [],
            'ltv_cac_ratio': [],
            'gross_margin_pct': [],
            'net_revenue_retention': [],
            'magic_number': [],
            'rule_of_40': [],
            'months_to_payback': [],
            'gross_profit': [],
            'ebitda': [],
            'ebitda_margin': []
        }

        for month in range(self.periods):
            # Total customers
            total_customers = sum([
                revenue['customers_by_tier']['hosts_premium'][month],
                revenue['customers_by_tier']['hosts_pro'][month],
                revenue['customers_by_tier']['renters_active'][month]
            ])

            # ARPU
            arpu = revenue['total_mrr'][month] / total_customers if total_customers > 0 else 0

            # CAC
            cac = (expenses['sales_marketing'][month] / expenses['new_customers'][month]) if expenses['new_customers'][month] > 0 else 0

            # Gross Margin
            gross_profit = revenue['total_mrr'][month] - expenses['cogs'][month]
            gross_margin_pct = (gross_profit / revenue['total_mrr'][month]) if revenue['total_mrr'][month] > 0 else 0

            # LTV (simplified: ARPU * Gross Margin / Churn Rate)
            avg_churn = (self.assumptions['churn_rate_premium'] + self.assumptions['churn_rate_pro'] + self.assumptions['churn_rate_renter']) / 3
            ltv = (arpu * gross_margin_pct / avg_churn) if avg_churn > 0 else 0

            # LTV:CAC Ratio
            ltv_cac = (ltv / cac) if cac > 0 else 0

            # NRR (simplified month-over-month comparison)
            if month > 0:
                nrr = (revenue['total_mrr'][month] / revenue['total_mrr'][month-1]) if revenue['total_mrr'][month-1] > 0 else 1
            else:
                nrr = 1.0

            # Magic Number
            if month >= 3:
                new_arr = (revenue['total_mrr'][month] - revenue['total_mrr'][month-3]) * 12
                sm_spend = sum(expenses['sales_marketing'][month-3:month])
                magic_number = (new_arr / sm_spend) if sm_spend > 0 else 0
            else:
                magic_number = 0

            # EBITDA
            ebitda = revenue['total_mrr'][month] - expenses['total_opex'][month]
            ebitda_margin = (ebitda / revenue['total_mrr'][month]) if revenue['total_mrr'][month] > 0 else 0

            # Growth rate (YoY)
            if month >= 12:
                growth_rate = ((revenue['total_mrr'][month] - revenue['total_mrr'][month-12]) / revenue['total_mrr'][month-12]) if revenue['total_mrr'][month-12] > 0 else 0
            else:
                growth_rate = 0

            # Rule of 40
            rule_of_40 = (growth_rate * 100) + (ebitda_margin * 100)

            # Months to Payback CAC
            months_to_payback = (cac / (arpu * gross_margin_pct)) if (arpu * gross_margin_pct) > 0 else 0

            metrics['customers_total'].append(int(total_customers))
            metrics['arpu'].append(arpu)
            metrics['cac'].append(cac)
            metrics['ltv'].append(ltv)
            metrics['ltv_cac_ratio'].append(ltv_cac)
            metrics['gross_margin_pct'].append(gross_margin_pct)
            metrics['net_revenue_retention'].append(nrr)
            metrics['magic_number'].append(magic_number)
            metrics['rule_of_40'].append(rule_of_40)
            metrics['months_to_payback'].append(months_to_payback)
            metrics['gross_profit'].append(gross_profit)
            metrics['ebitda'].append(ebitda)
            metrics['ebitda_margin'].append(ebitda_margin)

        self.results['metrics'] = metrics
        return metrics

    def calculate_break_even(self):
        """Calculate break-even analysis"""
        revenue = self.results['revenue']
        expenses = self.results['expenses']
        metrics = self.results['metrics']

        break_even = {
            'cash_flow_break_even_month': None,
            'ebitda_break_even_month': None,
            'contribution_margin_break_even_customers': None,
            'unit_economics_break_even_customers': None
        }

        # Find cash flow break-even (revenue >= opex)
        for month in range(self.periods):
            if revenue['total_mrr'][month] >= expenses['total_opex'][month]:
                break_even['cash_flow_break_even_month'] = month + 1
                break

        # Find EBITDA break-even
        for month in range(self.periods):
            if metrics['ebitda'][month] >= 0:
                break_even['ebitda_break_even_month'] = month + 1
                break

        # Contribution margin break-even customers
        # Fixed costs / Contribution margin per customer
        avg_fixed_costs = (
            self.assumptions['rd_min_monthly'] +
            self.assumptions['ga_min_monthly']
        )
        avg_arpu = sum(metrics['arpu'][:12]) / 12
        avg_variable_cost_per_customer = avg_arpu * self.assumptions['total_cogs_percent']
        contribution_margin = avg_arpu - avg_variable_cost_per_customer

        if contribution_margin > 0:
            break_even['contribution_margin_break_even_customers'] = int(avg_fixed_costs / contribution_margin)

        # Unit economics break-even (LTV >= 3x CAC)
        for month in range(self.periods):
            if metrics['ltv_cac_ratio'][month] >= 3.0:
                break_even['unit_economics_break_even_customers'] = metrics['customers_total'][month]
                break

        self.results['break_even'] = break_even
        return break_even

    def run_full_model(self):
        """Execute all calculations in sequence"""
        print("[*] Calculating revenue model...")
        self.calculate_revenue_model()

        print("[*] Calculating expenses...")
        self.calculate_expenses()

        print("[*] Calculating cash flow...")
        self.calculate_cash_flow()

        print("[*] Calculating SaaS metrics...")
        self.calculate_metrics()

        print("[*] Calculating break-even analysis...")
        self.calculate_break_even()

        print("[OK] All calculations complete!\n")
        return self.results

# =============================================================================
# EXCEL GENERATOR
# =============================================================================

class ExcelGenerator:
    def __init__(self, model_results, assumptions):
        self.results = model_results
        self.assumptions = assumptions
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

    def create_assumptions_sheet(self):
        """Create input assumptions sheet"""
        ws = self.wb.create_sheet("Assumptions")

        # Title
        ws['A1'] = f"{self.assumptions['company_name']} - Financial Assumptions"
        ws['A1'].font = ExcelStyles.TITLE_FONT
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Model Date: {self.assumptions['model_date']}"
        ws['A2'].font = Font(italic=True)

        row = 4

        # Pricing Tiers
        ws[f'A{row}'] = "PRICING TIERS"
        ws[f'A{row}'].font = ExcelStyles.SUBHEADER_FONT
        ws[f'A{row}'].fill = ExcelStyles.HEADER_FILL
        row += 1

        for tier_key, tier_data in self.assumptions['pricing_tiers'].items():
            ws[f'A{row}'] = tier_data['name']
            ws[f'B{row}'] = tier_data['price_monthly']
            ws[f'B{row}'].number_format = '$#,##0.00'
            ws[f'C{row}'] = tier_data['commission_rate']
            ws[f'C{row}'].number_format = '0.0%'
            ws[f'D{row}'] = tier_data['description']
            row += 1

        row += 2

        # Growth Assumptions
        ws[f'A{row}'] = "GROWTH ASSUMPTIONS"
        ws[f'A{row}'].font = ExcelStyles.SUBHEADER_FONT
        ws[f'A{row}'].fill = ExcelStyles.HEADER_FILL
        row += 1

        growth_items = [
            ('Monthly Growth Rate Y1', self.assumptions['monthly_growth_rate_y1'], '0.0%'),
            ('Monthly Growth Rate Y2', self.assumptions['monthly_growth_rate_y2'], '0.0%'),
            ('Monthly Growth Rate Y3', self.assumptions['monthly_growth_rate_y3'], '0.0%'),
            ('Monthly Growth Rate Y4-5', self.assumptions['monthly_growth_rate_y4'], '0.0%'),
            ('Free to Premium Conversion', self.assumptions['free_to_premium_rate'], '0.0%'),
            ('Premium to Pro Conversion', self.assumptions['premium_to_pro_rate'], '0.0%'),
        ]

        for label, value, fmt in growth_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = fmt
            ws[f'B{row}'].fill = ExcelStyles.INPUT_FILL
            row += 1

        row += 2

        # Churn Rates
        ws[f'A{row}'] = "CHURN RATES (Monthly)"
        ws[f'A{row}'].font = ExcelStyles.SUBHEADER_FONT
        ws[f'A{row}'].fill = ExcelStyles.HEADER_FILL
        row += 1

        churn_items = [
            ('Free Hosts', self.assumptions['churn_rate_free'], '0.0%'),
            ('Premium Hosts', self.assumptions['churn_rate_premium'], '0.0%'),
            ('Pro Hosts', self.assumptions['churn_rate_pro'], '0.0%'),
            ('Renters', self.assumptions['churn_rate_renter'], '0.0%'),
        ]

        for label, value, fmt in churn_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = fmt
            ws[f'B{row}'].fill = ExcelStyles.INPUT_FILL
            row += 1

        row += 2

        # CAC by Channel
        ws[f'A{row}'] = "CAC BY CHANNEL"
        ws[f'A{row}'].font = ExcelStyles.SUBHEADER_FONT
        ws[f'A{row}'].fill = ExcelStyles.HEADER_FILL
        row += 1

        for channel, cac in self.assumptions['cac_by_channel'].items():
            ws[f'A{row}'] = channel.replace('_', ' ').title()
            ws[f'B{row}'] = cac
            ws[f'B{row}'].number_format = '$#,##0'
            ws[f'C{row}'] = self.assumptions['channel_mix'].get(channel, 0)
            ws[f'C{row}'].number_format = '0.0%'
            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40

        print("[OK] Assumptions sheet created")

    def create_revenue_sheet(self):
        """Create detailed revenue model sheet"""
        ws = self.wb.create_sheet("Revenue Model")

        # Headers
        ws['A1'] = "Revenue Model - Monthly Detail"
        ws['A1'].font = ExcelStyles.TITLE_FONT
        ws.merge_cells('A1:H1')

        # Column headers
        headers = [
            'Month', 'Year', 'Host Commission', 'Renter Fees',
            'Premium Subs', 'Pro Subs', 'Expansion', 'Total MRR', 'ARR'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = ExcelStyles.HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL
            cell.alignment = ExcelStyles.CENTER

        # Data
        revenue = self.results['revenue']
        for month_idx in range(len(revenue['months'])):
            row = month_idx + 4
            ws.cell(row=row, column=1, value=month_idx + 1)
            ws.cell(row=row, column=2, value=f"Y{(month_idx // 12) + 1}")
            ws.cell(row=row, column=3, value=revenue['host_commission'][month_idx])
            ws.cell(row=row, column=4, value=revenue['renter_fees'][month_idx])
            ws.cell(row=row, column=5, value=revenue['premium_subscriptions'][month_idx])
            ws.cell(row=row, column=6, value=revenue['pro_subscriptions'][month_idx])
            ws.cell(row=row, column=7, value=revenue['expansion_revenue'][month_idx])
            ws.cell(row=row, column=8, value=revenue['total_mrr'][month_idx])
            ws.cell(row=row, column=9, value=revenue['arr'][month_idx])

            # Format as currency
            for col in range(3, 10):
                ws.cell(row=row, column=col).number_format = '$#,##0'

        # Add revenue fill to MRR/ARR columns
        for row in range(4, len(revenue['months']) + 4):
            ws.cell(row=row, column=8).fill = ExcelStyles.REVENUE_FILL
            ws.cell(row=row, column=9).fill = ExcelStyles.REVENUE_FILL

        # Format columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 14

        # Add chart
        chart = LineChart()
        chart.title = "Revenue Growth Over Time"
        chart.y_axis.title = "Revenue ($)"
        chart.x_axis.title = "Month"

        data = Reference(ws, min_col=8, min_row=3, max_row=len(revenue['months']) + 3)
        cats = Reference(ws, min_col=1, min_row=4, max_row=len(revenue['months']) + 3)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "K3")

        print("[OK] Revenue model sheet created")

    def create_expenses_sheet(self):
        """Create detailed expenses sheet"""
        ws = self.wb.create_sheet("Expenses")

        # Headers
        ws['A1'] = "Operating Expenses - Monthly Detail"
        ws['A1'].font = ExcelStyles.TITLE_FONT
        ws.merge_cells('A1:H1')

        headers = [
            'Month', 'Year', 'COGS', 'Sales & Marketing',
            'R&D', 'G&A', 'Total OpEx', 'Headcount', 'New Customers'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = ExcelStyles.HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL
            cell.alignment = ExcelStyles.CENTER

        # Data
        expenses = self.results['expenses']
        for month_idx in range(len(expenses['months'])):
            row = month_idx + 4
            ws.cell(row=row, column=1, value=month_idx + 1)
            ws.cell(row=row, column=2, value=f"Y{(month_idx // 12) + 1}")
            ws.cell(row=row, column=3, value=expenses['cogs'][month_idx])
            ws.cell(row=row, column=4, value=expenses['sales_marketing'][month_idx])
            ws.cell(row=row, column=5, value=expenses['research_development'][month_idx])
            ws.cell(row=row, column=6, value=expenses['general_admin'][month_idx])
            ws.cell(row=row, column=7, value=expenses['total_opex'][month_idx])
            ws.cell(row=row, column=8, value=expenses['headcount'][month_idx])
            ws.cell(row=row, column=9, value=expenses['new_customers'][month_idx])

            # Format as currency
            for col in range(3, 8):
                ws.cell(row=row, column=col).number_format = '$#,##0'

        # Add cost fill
        for row in range(4, len(expenses['months']) + 4):
            ws.cell(row=row, column=7).fill = ExcelStyles.COST_FILL

        # Format columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 14

        print("[OK] Expenses sheet created")

    def create_cash_flow_sheet(self):
        """Create cash flow statement"""
        ws = self.wb.create_sheet("Cash Flow")

        # Headers
        ws['A1'] = "Cash Flow Statement"
        ws['A1'].font = ExcelStyles.TITLE_FONT
        ws.merge_cells('A1:H1')

        headers = [
            'Month', 'Year', 'Revenue', 'Expenses', 'Net Income',
            'Funding Events', 'Cash Balance', 'Burn Rate', 'Runway (Months)'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = ExcelStyles.HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL
            cell.alignment = ExcelStyles.CENTER

        # Data
        cash = self.results['cash_flow']
        for month_idx in range(len(cash['months'])):
            row = month_idx + 4
            ws.cell(row=row, column=1, value=month_idx + 1)
            ws.cell(row=row, column=2, value=f"Y{(month_idx // 12) + 1}")
            ws.cell(row=row, column=3, value=cash['revenue'][month_idx])
            ws.cell(row=row, column=4, value=cash['total_expenses'][month_idx])
            ws.cell(row=row, column=5, value=cash['net_income'][month_idx])
            ws.cell(row=row, column=6, value=cash['funding_events'][month_idx])
            ws.cell(row=row, column=7, value=cash['cash_balance'][month_idx])
            ws.cell(row=row, column=8, value=cash['burn_rate'][month_idx])
            runway = cash['runway_months'][month_idx]
            ws.cell(row=row, column=9, value=runway if runway < 999 else 'Infinite')

            # Format as currency
            for col in [3, 4, 5, 6, 7, 8]:
                ws.cell(row=row, column=col).number_format = '$#,##0'

            # Conditional formatting for cash balance
            if cash['cash_balance'][month_idx] < 0:
                ws.cell(row=row, column=7).font = Font(color='FF0000', bold=True)

        # Format columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 14

        # Chart
        chart = LineChart()
        chart.title = "Cash Balance Over Time"
        chart.y_axis.title = "Cash ($)"
        chart.x_axis.title = "Month"

        data = Reference(ws, min_col=7, min_row=3, max_row=len(cash['months']) + 3)
        cats = Reference(ws, min_col=1, min_row=4, max_row=len(cash['months']) + 3)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "K3")

        print("[OK] Cash flow sheet created")

    def create_metrics_dashboard(self):
        """Create KPI dashboard"""
        ws = self.wb.create_sheet("Metrics Dashboard")

        # Title
        ws['A1'] = "SaaS Metrics Dashboard"
        ws['A1'].font = ExcelStyles.TITLE_FONT
        ws.merge_cells('A1:F1')

        metrics = self.results['metrics']

        # Key metrics summary (latest month)
        latest = len(metrics['months']) - 1

        row = 3
        ws[f'A{row}'] = "KEY METRICS (Latest Month)"
        ws[f'A{row}'].font = ExcelStyles.SUBHEADER_FONT
        ws[f'A{row}'].fill = ExcelStyles.METRIC_FILL
        ws.merge_cells(f'A{row}:B{row}')
        row += 1

        key_metrics = [
            ('ARR', metrics['arr'][latest], '$#,##0'),
            ('MRR', metrics['mrr'][latest], '$#,##0'),
            ('Total Customers', metrics['customers_total'][latest], '#,##0'),
            ('ARPU', metrics['arpu'][latest], '$#,##0'),
            ('CAC', metrics['cac'][latest], '$#,##0'),
            ('LTV', metrics['ltv'][latest], '$#,##0'),
            ('LTV:CAC Ratio', metrics['ltv_cac_ratio'][latest], '0.00'),
            ('Gross Margin %', metrics['gross_margin_pct'][latest], '0.0%'),
            ('NRR', metrics['net_revenue_retention'][latest], '0.0%'),
            ('Magic Number', metrics['magic_number'][latest], '0.00'),
            ('Rule of 40', metrics['rule_of_40'][latest], '0'),
            ('CAC Payback (Months)', metrics['months_to_payback'][latest], '0.0'),
        ]

        for label, value, fmt in key_metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = fmt
            ws[f'B{row}'].fill = ExcelStyles.METRIC_FILL
            ws[f'B{row}'].font = Font(bold=True, size=12)
            row += 1

        # Break-even analysis
        row += 2
        ws[f'A{row}'] = "BREAK-EVEN ANALYSIS"
        ws[f'A{row}'].font = ExcelStyles.SUBHEADER_FONT
        ws[f'A{row}'].fill = ExcelStyles.HEADER_FILL
        ws.merge_cells(f'A{row}:B{row}')
        row += 1

        break_even = self.results['break_even']
        be_metrics = [
            ('Cash Flow Break-Even Month', break_even['cash_flow_break_even_month'] or 'Not Yet', '0'),
            ('EBITDA Break-Even Month', break_even['ebitda_break_even_month'] or 'Not Yet', '0'),
            ('Break-Even Customers (Contribution)', break_even['contribution_margin_break_even_customers'] or 'N/A', '#,##0'),
            ('Unit Economics Break-Even Customers', break_even['unit_economics_break_even_customers'] or 'N/A', '#,##0'),
        ]

        for label, value, fmt in be_metrics:
            ws[f'A{row}'] = label
            if isinstance(value, (int, float)):
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = fmt
            else:
                ws[f'B{row}'] = value
            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

        print("[OK] Metrics dashboard created")

    def create_break_even_sheet(self):
        """Create detailed break-even analysis with visualization"""
        ws = self.wb.create_sheet("Break-Even Analysis")

        ws['A1'] = "Break-Even Analysis & Path to Profitability"
        ws['A1'].font = ExcelStyles.TITLE_FONT
        ws.merge_cells('A1:F1')

        # Headers for month-by-month comparison
        headers = ['Month', 'Year', 'Revenue', 'OpEx', 'Net Income', 'Cumulative Profit']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = ExcelStyles.HEADER_FONT
            cell.fill = ExcelStyles.HEADER_FILL

        # Data
        revenue = self.results['revenue']
        expenses = self.results['expenses']
        cash = self.results['cash_flow']

        cumulative_profit = 0
        for month_idx in range(len(revenue['months'])):
            row = month_idx + 4
            net = cash['net_income'][month_idx]
            cumulative_profit += net

            ws.cell(row=row, column=1, value=month_idx + 1)
            ws.cell(row=row, column=2, value=f"Y{(month_idx // 12) + 1}")
            ws.cell(row=row, column=3, value=revenue['total_mrr'][month_idx])
            ws.cell(row=row, column=4, value=expenses['total_opex'][month_idx])
            ws.cell(row=row, column=5, value=net)
            ws.cell(row=row, column=6, value=cumulative_profit)

            # Highlight break-even month
            if month_idx > 0 and cumulative_profit >= 0 and (cumulative_profit - net) < 0:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                    ws.cell(row=row, column=col).font = Font(bold=True)

            # Format currency
            for col in [3, 4, 5, 6]:
                ws.cell(row=row, column=col).number_format = '$#,##0'

        # Chart
        chart = LineChart()
        chart.title = "Path to Break-Even"
        chart.y_axis.title = "Amount ($)"
        chart.x_axis.title = "Month"

        # Revenue line
        revenue_data = Reference(ws, min_col=3, min_row=3, max_row=len(revenue['months']) + 3)
        # OpEx line
        opex_data = Reference(ws, min_col=4, min_row=3, max_row=len(revenue['months']) + 3)
        # Cumulative profit line
        profit_data = Reference(ws, min_col=6, min_row=3, max_row=len(revenue['months']) + 3)

        cats = Reference(ws, min_col=1, min_row=4, max_row=len(revenue['months']) + 3)

        chart.add_data(revenue_data, titles_from_data=True)
        chart.add_data(opex_data, titles_from_data=True)
        chart.add_data(profit_data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "H3")

        print("[OK] Break-even analysis sheet created")

    def create_executive_summary(self):
        """Create executive summary sheet"""
        ws = self.wb.create_sheet("Executive Summary", 0)  # Make first sheet

        # Title
        ws['A1'] = f"{self.assumptions['company_name']}"
        ws['A1'].font = Font(name='Calibri', size=24, bold=True, color='1F4E78')
        ws.merge_cells('A1:D1')

        ws['A2'] = "5-Year Financial Projection"
        ws['A2'].font = Font(name='Calibri', size=16, italic=True)
        ws.merge_cells('A2:D2')

        ws['A3'] = f"Generated: {self.assumptions['model_date']}"
        ws['A3'].font = Font(italic=True, size=10)

        # Key highlights
        row = 5
        ws[f'A{row}'] = "KEY HIGHLIGHTS"
        ws[f'A{row}'].font = ExcelStyles.SUBHEADER_FONT
        ws[f'A{row}'].fill = ExcelStyles.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        metrics = self.results['metrics']
        revenue = self.results['revenue']
        cash = self.results['cash_flow']

        # Year 1 vs Year 5 metrics
        y1_end = 11  # Month 12 (index 11)
        y5_end = 59  # Month 60 (index 59)

        highlights = [
            ('ARR (Year 1 → Year 5)',
             f"${revenue['arr'][y1_end]:,.0f} → ${revenue['arr'][y5_end]:,.0f}",
             f"{((revenue['arr'][y5_end] / revenue['arr'][y1_end]) - 1) * 100:.0f}% growth"),

            ('Total Customers (Year 1 → Year 5)',
             f"{metrics['customers_total'][y1_end]:,} → {metrics['customers_total'][y5_end]:,}",
             f"{((metrics['customers_total'][y5_end] / metrics['customers_total'][y1_end]) - 1) * 100:.0f}% growth"),

            ('LTV:CAC Ratio (Year 5)',
             f"{metrics['ltv_cac_ratio'][y5_end]:.1f}:1",
             "Strong unit economics" if metrics['ltv_cac_ratio'][y5_end] >= 3 else "Needs improvement"),

            ('Gross Margin (Year 5)',
             f"{metrics['gross_margin_pct'][y5_end] * 100:.0f}%",
             "Healthy margins" if metrics['gross_margin_pct'][y5_end] >= 0.70 else "Below target"),

            ('Cash Position (End of Year 5)',
             f"${cash['cash_balance'][y5_end]:,.0f}",
             "Strong" if cash['cash_balance'][y5_end] > 0 else "Requires funding"),
        ]

        for metric, value, note in highlights:
            ws[f'A{row}'] = metric
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(size=12, bold=True)
            ws[f'C{row}'] = note
            ws[f'C{row}'].font = Font(italic=True, color='666666')
            row += 1

        row += 2

        # Break-even summary
        ws[f'A{row}'] = "BREAK-EVEN MILESTONES"
        ws[f'A{row}'].font = ExcelStyles.SUBHEADER_FONT
        ws[f'A{row}'].fill = ExcelStyles.HEADER_FILL
        ws.merge_cells(f'A{row}:D{row}')
        row += 2

        break_even = self.results['break_even']

        be_summary = [
            ('Cash Flow Break-Even', break_even['cash_flow_break_even_month'] or 'Not in projection'),
            ('EBITDA Positive', break_even['ebitda_break_even_month'] or 'Not in projection'),
            ('Unit Economics Validated', f"Month {break_even.get('unit_economics_break_even_customers', 'TBD')}" if break_even.get('unit_economics_break_even_customers') else 'Not in projection'),
        ]

        for label, value in be_summary:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = f"Month {value}" if isinstance(value, int) else value
            ws[f'B{row}'].font = Font(bold=True)
            row += 1

        # Format
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20

        print("[OK] Executive summary created")

    def generate_excel(self, filename='saas_financial_model.xlsx'):
        """Generate complete Excel file"""
        print("\n" + "="*60)
        print("GENERATING EXCEL FINANCIAL MODEL")
        print("="*60 + "\n")

        self.create_executive_summary()
        self.create_assumptions_sheet()
        self.create_revenue_sheet()
        self.create_expenses_sheet()
        self.create_cash_flow_sheet()
        self.create_metrics_dashboard()
        self.create_break_even_sheet()

        # Save
        self.wb.save(filename)
        print(f"\n[OK] Financial model saved to: {filename}")
        print(f"[*] Full path: {os.path.abspath(filename)}\n")

        return filename

# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    print("\n" + "="*60)
    print("MARKET SPACE - FINANCIAL MODEL GENERATOR")
    print("="*60 + "\n")

    # Try to load business model canvas
    canvas_data = parse_business_model_canvas()

    # If canvas has data, merge with assumptions
    if canvas_data:
        ASSUMPTIONS.update(canvas_data)

    # Create financial model
    print("[*] Initializing financial model...\n")
    model = SaaSFinancialModel(ASSUMPTIONS)

    # Run calculations
    results = model.run_full_model()

    # Generate Excel
    generator = ExcelGenerator(results, ASSUMPTIONS)
    output_file = generator.generate_excel('saas_financial_model.xlsx')

    # Print summary
    print("="*60)
    print("MODEL SUMMARY")
    print("="*60)
    print(f"Company: {ASSUMPTIONS['company_name']}")
    print(f"Projection Period: 60 months (5 years)")
    print(f"Starting Cash: ${ASSUMPTIONS['starting_cash']:,.0f}")
    print(f"\nYear 1 ARR: ${results['revenue']['arr'][11]:,.0f}")
    print(f"Year 5 ARR: ${results['revenue']['arr'][59]:,.0f}")
    print(f"Total Growth: {((results['revenue']['arr'][59] / results['revenue']['arr'][11]) - 1) * 100:.0f}%")

    if results['break_even']['cash_flow_break_even_month']:
        print(f"\n[OK] Cash Flow Break-Even: Month {results['break_even']['cash_flow_break_even_month']}")
    else:
        print(f"\n[!] Cash Flow Break-Even: Not achieved in 5-year projection")

    print("\n" + "="*60)
    print("COMPLETE! Open the Excel file to view detailed projections.")
    print("="*60 + "\n")

if __name__ == "__main__":
    main()
